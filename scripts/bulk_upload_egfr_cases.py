#!/usr/bin/env python3
import argparse
import json
import math
import mimetypes
import os
import sys
import time
from datetime import datetime, timezone
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import load_workbook


DEFAULT_XLSX = Path("data/Case041_to_Case100_CKD_USG_Extracted.xlsx")
DEFAULT_ZIP_DIR = Path("data/7. US images with Laboratory reports (Case041 to Case100)")
DEFAULT_HOSPITAL_ID = "MIL-NDL-DL"
DEFAULT_HOSPITAL_NAME = "Mahajan Imaging & Labs, New Delhi"
DEFAULT_USERNAME = "MIL-NDL-DL"
DEFAULT_START = 57
DEFAULT_END = 100


def request_json(url, *, method="GET", token=None, body=None, data=None, content_type="application/json"):
  headers = {}
  payload = data
  if body is not None:
    payload = json.dumps(body).encode("utf-8")
    headers["Content-Type"] = content_type
  elif data is not None and content_type:
    headers["Content-Type"] = content_type
  if token:
    headers["Authorization"] = f"Bearer {token}"
  req = Request(url, data=payload, headers=headers, method=method)
  try:
    with urlopen(req, timeout=120) as response:
      raw = response.read()
      if not raw:
        return {}
      return json.loads(raw.decode("utf-8"))
  except HTTPError as err:
    details = err.read().decode("utf-8", errors="replace")
    try:
      parsed = json.loads(details)
      message = parsed.get("error") or details
    except json.JSONDecodeError:
      message = details
    raise RuntimeError(f"{method} {url} failed ({err.code}): {message}") from err
  except URLError as err:
    raise RuntimeError(f"{method} {url} failed: {err.reason}") from err


def clean_cell(value):
  if value is None:
    return "N/A"
  text = str(value).strip()
  return text if text else "N/A"


def yes_no_na(value):
  text = clean_cell(value).lower()
  if text in {"yes", "y", "true", "1"}:
    return "Yes"
  if text in {"no", "n", "false", "0"}:
    return "No"
  return "N/A"


def yes_no_required(value, label):
  result = yes_no_na(value)
  if result == "N/A":
    raise ValueError(f"{label} must be Yes or No; N/A is not allowed.")
  return result


def sex_value(value):
  text = clean_cell(value).lower()
  if text.startswith("m"):
    return "Male"
  if text.startswith("f"):
    return "Female"
  return "Other"


def number_or_dash(value):
  text = clean_cell(value)
  if text == "N/A":
    return "-"
  try:
    number = float(text)
  except ValueError:
    return "-"
  if not math.isfinite(number) or number < 0:
    return "-"
  if number.is_integer():
    return str(int(number))
  return str(round(number, 2)).rstrip("0").rstrip(".")


def optional_text(value):
  text = clean_cell(value)
  return "-" if text == "N/A" else text


def echogenicity_value(value):
  text = clean_cell(value)
  mapping = {
    "Normal": "Normal",
    "Mild ↑": "Mild Increased",
    "Mild Increased": "Mild Increased",
    "Moderate Increased": "Moderate Increased",
    "Severe Increased": "Severe Increased",
    "Increased (not specified)": "Increased (Not specified)",
    "Increased (Not specified)": "Increased (Not specified)",
    "N/A": "N/A"
  }
  return mapping.get(text, "N/A")


def kidney_size_value(value):
  text = clean_cell(value)
  return text if text in {"Normal", "Small", "Enlarged", "N/A"} else "N/A"


def parenchymal_texture_value(value):
  text = clean_cell(value)
  return text if text in {"Normal", "Altered", "N/A"} else "N/A"


def case_id(number):
  return f"Case{number:03d}"


def load_rows(xlsx_path):
  wb = load_workbook(xlsx_path, read_only=True, data_only=True)
  ws = wb["Extracted_Data"]
  headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
  rows = {}
  for values in ws.iter_rows(min_row=2, values_only=True):
    row = dict(zip(headers, values))
    case_no = clean_cell(row.get("Case No."))
    if case_no != "N/A":
      rows[case_no] = row
  return rows


def kidney_payload(row, side):
  prefix = f"{side} Kidney"
  cortical_prefix = f"{side} Cortical/Parenchymal Thickness"
  return {
    "lengthCm": number_or_dash(row.get(f"{prefix} Length (cm)")),
    "widthCm": number_or_dash(row.get(f"{prefix} Width (cm)")),
    "corticalThicknessMm": number_or_dash(row.get(f"{cortical_prefix} (mm)")),
    "echogenicity": echogenicity_value(row.get(f"{side} Echogenicity")),
    "structural": {
      "kidneySize": kidney_size_value(row.get(f"{prefix} Size")),
      "parenchymalTexture": parenchymal_texture_value(row.get(f"{side} Parenchymal Texture")),
      "cysts": yes_no_na(row.get(f"{side} Cysts")),
      "stones": yes_no_na(row.get(f"{side} Stones")),
      "hydronephrosis": yes_no_na(row.get(f"{side} Hydronephrosis")),
      "others": optional_text(row.get(f"{side} Others"))
    }
  }


def submission_payload(row, zip_path, hospital_id, hospital_name):
  known_ckd = yes_no_required(row.get("CKD (Yes/No)"), "Chronic Kidney Disease (CKD)")
  if known_ckd == "No":
    ckd_stage = "Normal"
  else:
    ckd_stage = "Other"

  file_type = mimetypes.guess_type(zip_path.name)[0] or "application/zip"
  return {
    "hospitalId": hospital_id,
    "hospitalName": hospital_name,
    "hospitalSessionId": hospital_id,
    "hospitalSessionName": hospital_name,
    "consentId": None,
    "studyFlow": "egfr",
    "studyId": None,
    "enrollmentDate": "-",
    "siteCenter": "-",
    "consentObtained": "-",
    "uploadMode": "package",
    "uhid": clean_cell(row.get("Case No.")),
    "age": number_or_dash(row.get("Age")),
    "sex": sex_value(row.get("Sex")),
    "heightCm": "-",
    "weight": "-",
    "bmi": "-",
    "ethnicity": "-",
    "occupation": "-",
    "knownCkd": known_ckd,
    "ckdDuration": "-",
    "ckdStage": ckd_stage,
    "ckdStageRemarks": "CKD reported; stage not collected in source sheet." if known_ckd == "Yes" else "-",
    "dialysis": "-",
    "dialysisFrequency": "-",
    "diabetic": yes_no_na(row.get("Diabetes")),
    "diabeticStage": "-",
    "diabetesDuration": "-",
    "hypertension": yes_no_na(row.get("Hypertension")),
    "hypertensionDuration": "-",
    "cardiovascularDisease": "-",
    "familyKidneyHistory": "-",
    "ultrasoundFindings": {
      "right": kidney_payload(row, "Right"),
      "left": kidney_payload(row, "Left"),
      "imageQuality": {
        "adequateForAnalysis": "N/A"
      },
      "annotationDetails": {
        "kidneyBoundingPointsDetected": "N/A"
      }
    },
    "kfreForm": None,
    "files": [{
      "fieldName": "patientPackage",
      "name": zip_path.name,
      "type": file_type,
      "size": zip_path.stat().st_size
    }]
  }


def login(base_url, username, password):
  response = request_json(f"{base_url}/api/auth/login", method="POST", body={
    "username": username,
    "password": password
  })
  token = response.get("token")
  if not token:
    raise RuntimeError("Login succeeded but no token was returned.")
  user = response.get("user") or {}
  return token, user


def upload_record(base_url, token, submission, zip_path):
  init = request_json(f"{base_url}/api/uploads/init", method="POST", token=token, body={
    "timestamp": datetime.now(timezone.utc).isoformat(),
    "submission": submission
  })
  upload_id = init["uploadId"]
  file_info = init["files"][0]
  chunk_size = int(init.get("chunkSize") or file_info.get("chunkSize") or 4 * 1024 * 1024)
  with zip_path.open("rb") as stream:
    chunk_index = 0
    while True:
      chunk = stream.read(chunk_size)
      if not chunk:
        break
      request_json(
        f"{base_url}/api/uploads/{upload_id}/files/0/chunks/{chunk_index}",
        method="PUT",
        token=token,
        data=chunk,
        content_type="application/octet-stream"
      )
      chunk_index += 1
  return request_json(f"{base_url}/api/uploads/{upload_id}/complete", method="POST", token=token, body={})


def parse_args():
  parser = argparse.ArgumentParser(description="Bulk upload eGFR Case057-Case100 records through the TANUH portal API.")
  parser.add_argument("--base-url", default=os.environ.get("PORTAL_BASE_URL", "http://127.0.0.1:8000"))
  parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
  parser.add_argument("--zip-dir", type=Path, default=DEFAULT_ZIP_DIR)
  parser.add_argument("--start", type=int, default=DEFAULT_START)
  parser.add_argument("--end", type=int, default=DEFAULT_END)
  parser.add_argument("--username", default=os.environ.get("BULK_UPLOAD_USERNAME", DEFAULT_USERNAME))
  parser.add_argument("--password", default=os.environ.get("BULK_UPLOAD_PASSWORD", ""))
  parser.add_argument("--hospital-id", default=os.environ.get("BULK_UPLOAD_HOSPITAL_ID", DEFAULT_HOSPITAL_ID))
  parser.add_argument("--hospital-name", default=os.environ.get("BULK_UPLOAD_HOSPITAL_NAME", DEFAULT_HOSPITAL_NAME))
  parser.add_argument("--execute", action="store_true", help="Actually upload records. Without this, only validates mappings.")
  parser.add_argument("--limit", type=int, default=0, help="Optional cap for test uploads.")
  return parser.parse_args()


def main():
  args = parse_args()
  rows = load_rows(args.xlsx)
  selected = []
  errors = []
  for number in range(args.start, args.end + 1):
    cid = case_id(number)
    row = rows.get(cid)
    zip_path = args.zip_dir / f"{cid}.zip"
    if not row:
      errors.append(f"{cid}: missing spreadsheet row")
      continue
    if not zip_path.exists():
      errors.append(f"{cid}: missing ZIP package at {zip_path}")
      continue
    selected.append((cid, row, zip_path))

  if args.limit:
    selected = selected[:args.limit]

  mapped = []
  valid_selected = []
  for cid, row, zip_path in selected:
    try:
      mapped.append(submission_payload(row, zip_path, args.hospital_id, args.hospital_name))
      valid_selected.append((cid, row, zip_path))
    except ValueError as err:
      errors.append(f"{cid}: {err}")
  selected = valid_selected
  report = {
    "mode": "execute" if args.execute else "dry-run",
    "baseUrl": args.base_url,
    "hospitalId": args.hospital_id,
    "caseRange": f"{case_id(args.start)}-{case_id(args.end)}",
    "recordsReady": len(mapped),
    "errors": errors,
    "cases": [
      {
        "case": payload["uhid"],
        "age": payload["age"],
        "sex": payload["sex"],
        "knownCkd": payload["knownCkd"],
        "diabetes": payload["diabetic"],
        "hypertension": payload["hypertension"],
        "zip": zip_path.name,
        "size": zip_path.stat().st_size
      }
      for payload, (_, _, zip_path) in zip(mapped, selected)
    ],
    "uploads": []
  }

  if errors:
    print(json.dumps(report, indent=2))
    return 2

  if not args.execute:
    print(json.dumps(report, indent=2))
    return 0

  if not args.password:
    print("Set BULK_UPLOAD_PASSWORD or pass --password before using --execute.", file=sys.stderr)
    return 2

  token, user = login(args.base_url.rstrip("/"), args.username, args.password)
  hospital_name = user.get("hospitalName") or args.hospital_name
  hospital_id = user.get("hospitalId") or args.hospital_id

  for cid, row, zip_path in selected:
    payload = submission_payload(row, zip_path, hospital_id, hospital_name)
    started = time.time()
    result = upload_record(args.base_url.rstrip("/"), token, payload, zip_path)
    report["uploads"].append({
      "case": cid,
      "ok": bool(result.get("ok")),
      "batchId": result.get("batchId"),
      "gcsSynced": result.get("gcsSynced"),
      "dbSaved": result.get("dbSaved"),
      "seconds": round(time.time() - started, 2)
    })
    print(f"{cid}: uploaded batch={result.get('batchId')} gcs={result.get('gcsSynced')} db={result.get('dbSaved')}")

  reports_dir = Path("data/reports")
  reports_dir.mkdir(parents=True, exist_ok=True)
  report_path = reports_dir / f"bulk_upload_egfr_{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}.json"
  report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
  print(f"Report written to {report_path}")
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
